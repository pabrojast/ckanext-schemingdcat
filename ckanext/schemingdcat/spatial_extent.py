"""
Spatial extent extraction utilities for geospatial files.

This module provides functionality to extract spatial extent (bounding box)
from various geospatial file formats including Shapefiles, GeoTIFF, and others.

IMPORTANT: This functionality is designed to work ONLY through the web interface
for form auto-fill purposes. It does NOT interfere with CKAN's API operations
(resource_create, package_create, etc.) as it only runs client-side via JavaScript
when users upload files through the web form.

API Safety:
- No hooks into CKAN's core upload/create actions
- Only activated via frontend JavaScript in upload.html
- Completely separate endpoint (/api/extract-spatial-extent)
- Graceful degradation when spatial libraries are not available
- Silent failure mode to avoid disrupting normal workflows
"""

import json
import logging
import tempfile
import os
import zipfile
import hashlib
import mimetypes
from datetime import datetime
from typing import Optional, Dict, Any, Tuple
from collections import Counter

log = logging.getLogger(__name__)

try:
    import fiona
    from fiona.crs import from_epsg
    FIONA_AVAILABLE = True
except ImportError:
    FIONA_AVAILABLE = False
    log.warning("Fiona not available - Shapefile extent extraction disabled")

try:
    import rasterio
    from rasterio.warp import transform_bounds
    RASTERIO_AVAILABLE = True
except ImportError:
    RASTERIO_AVAILABLE = False
    log.warning("Rasterio not available - Raster extent extraction disabled")

try:
    from pyproj import Transformer
    PYPROJ_AVAILABLE = True
except ImportError:
    PYPROJ_AVAILABLE = False
    log.warning("PyProj not available - CRS transformation limited")

try:
    import geopandas as gpd
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    try:
        import pypdf
        PDF_AVAILABLE = True
        PyPDF2 = pypdf  # Use pypdf as PyPDF2 replacement
    except ImportError:
        PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class SpatialExtentExtractor:
    """Extract spatial extent from geospatial files."""
    
    SUPPORTED_EXTENSIONS = {
        'shp': 'shapefile',
        'zip': 'zip_shapefile',  # ZIP files containing shapefiles
        'tif': 'geotiff',
        'tiff': 'geotiff',
        'geotiff': 'geotiff',
        'kml': 'kml',
        'gpkg': 'geopackage',
        'geojson': 'geojson',
        'json': 'geojson'  # Assume GeoJSON if JSON
    }
    
    def __init__(self):
        self.available_handlers = self._check_available_handlers()
    
    def _is_potential_spatial_file(self, file_path_or_url):
        """
        Check if a file (by path or URL) could potentially be a spatial file.
        This includes ZIP files that might contain shapefiles.
        """
        # Extract filename from URL or path
        if file_path_or_url.startswith(('http://', 'https://')):
            filename = file_path_or_url.split('/')[-1].split('?')[0]
        else:
            filename = os.path.basename(file_path_or_url)
        
        ext = self._get_file_extension(filename)
        
        # Direct spatial file extensions
        spatial_extensions = ['shp', 'tif', 'tiff', 'geotiff', 'kml', 'gpkg', 'geojson', 'json']
        if ext in spatial_extensions:
            return True
        
        # ZIP files are potential spatial files if they might contain shapefiles
        # We consider ALL ZIP files as potential spatial files for this check
        if ext == 'zip':
            return True
        
        return False
    
    def _check_available_handlers(self) -> Dict[str, bool]:
        """Check which file format handlers are available."""
        return {
            'shapefile': FIONA_AVAILABLE,
            'zip_shapefile': FIONA_AVAILABLE,
            'geotiff': RASTERIO_AVAILABLE,
            'kml': FIONA_AVAILABLE,
            'geopackage': FIONA_AVAILABLE,
            'geojson': FIONA_AVAILABLE
        }
    
    def _is_shapefile_zip(self, file_path: str) -> bool:
        """Check if a ZIP file contains a shapefile."""
        try:
            # First check if file exists and has size
            if not os.path.exists(file_path):
                log.debug(f"ZIP file does not exist: {file_path}")
                return False
            
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                log.debug(f"ZIP file has zero size: {file_path}")
                return False
                
            log.info(f"Checking ZIP contents for shapefile components in: {file_path} (size: {file_size} bytes)")
            
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                log.info(f"ZIP contains {len(file_list)} files")
                
                # Check if we have the essential shapefile components
                has_shp = any(f.lower().endswith('.shp') for f in file_list)
                has_shx = any(f.lower().endswith('.shx') for f in file_list)
                has_dbf = any(f.lower().endswith('.dbf') for f in file_list)
                
                # Log detailed information
                log.info(f"Shapefile components check - .shp: {has_shp}, .shx: {has_shx}, .dbf: {has_dbf}")
                
                # A valid shapefile ZIP should have at least .shp, .shx, and .dbf
                return has_shp and has_shx and has_dbf
                
        except zipfile.BadZipFile as e:
            log.info(f"Bad ZIP file {file_path}: {str(e)}")
            return False
        except Exception as e:
            log.info(f"Error checking ZIP contents {file_path}: {str(e)}")
            return False
    
    def can_extract_extent(self, file_path: str, trust_extension: bool = False) -> bool:
        """
        Check if extent can be extracted from the given file.
        
        Args:
            file_path: Path to the file to check
            trust_extension: If True, trust the file extension without further validation
                            (useful for uploaded files that may not be accessible yet)
        """
        ext = self._get_file_extension(file_path)
        if ext not in self.SUPPORTED_EXTENSIONS:
            log.info(f"Unsupported extension: {ext}")
            return False
        
        format_type = self.SUPPORTED_EXTENSIONS[ext]
        log.info(f"Format type for {ext}: {format_type}")
        
        # If we're trusting the extension (for uploaded files), skip validation
        if trust_extension:
            log.info(f"Trusting file extension without validation: {ext}")
            return self.available_handlers.get(format_type, False)
        
        # Special handling for ZIP files - check if they contain shapefiles
        if format_type == 'zip_shapefile':
            handler_available = self.available_handlers.get(format_type, False)
            is_shapefile_zip = self._is_shapefile_zip(file_path) if handler_available else False
            log.info(f"ZIP file validation - handler available: {handler_available}, is shapefile ZIP: {is_shapefile_zip}")
            return handler_available and is_shapefile_zip
        
        return self.available_handlers.get(format_type, False)
    
    def _get_file_extension(self, file_path: str) -> str:
        """Get file extension in lowercase."""
        return os.path.splitext(file_path)[1].lower().lstrip('.')
    
    def extract_extent(self, file_path: str, trust_extension: bool = False) -> Optional[Dict[str, Any]]:
        """
        Extract spatial extent from a geospatial file.
        
        Args:
            file_path: Path to the geospatial file
            trust_extension: If True, trust the file extension without further validation
                            (useful for uploaded files that may not be accessible yet)
            
        Returns:
            GeoJSON Polygon representing the extent in WGS84, or None if extraction fails
        """
        try:
            log.info(f"Starting extent extraction for: {file_path} (trust_extension: {trust_extension})")
            
            if not os.path.exists(file_path):
                log.info(f"File not found: {file_path}")
                return None
            
            file_size = os.path.getsize(file_path)
            log.info(f"File size: {file_size} bytes")
            
            if file_size == 0:
                log.info(f"File has zero size: {file_path}")
                return None
            
            can_extract = self.can_extract_extent(file_path, trust_extension)
            log.info(f"Can extract extent: {can_extract}")
            
            if not can_extract:
                log.info(f"Cannot extract extent from file: {file_path}")
                ext = self._get_file_extension(file_path)
                log.info(f"File extension: {ext}")
                log.info(f"Supported extensions: {list(self.SUPPORTED_EXTENSIONS.keys())}")
                if ext in self.SUPPORTED_EXTENSIONS:
                    format_type = self.SUPPORTED_EXTENSIONS[ext]
                    log.info(f"Format type: {format_type}")
                    log.info(f"Handler available: {self.available_handlers.get(format_type, False)}")
                return None
            
            ext = self._get_file_extension(file_path)
            format_type = self.SUPPORTED_EXTENSIONS[ext]
            log.info(f"Processing as format type: {format_type}")
            
            if format_type == 'shapefile':
                return self._extract_shapefile_extent(file_path)
            elif format_type == 'zip_shapefile':
                return self._extract_zip_shapefile_extent(file_path)
            elif format_type == 'geotiff':
                return self._extract_raster_extent(file_path)
            elif format_type in ['kml', 'geopackage', 'geojson']:
                return self._extract_vector_extent(file_path)
            else:
                log.info(f"Unsupported format type: {format_type}")
                return None
                
        except Exception as e:
            # Use info level for debugging the 400 error
            log.error(f"Error extracting extent from {file_path}: {str(e)}", exc_info=True)
            return None
    
    def _extract_shapefile_extent(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract extent from Shapefile."""
        log.info(f"Extracting extent from shapefile: {file_path}")
        
        if not FIONA_AVAILABLE:
            log.info("Fiona not available for shapefile extraction")
            return None
        
        try:
            log.info(f"Opening shapefile with fiona: {file_path}")
            with fiona.open(file_path) as src:
                log.info(f"Shapefile opened successfully - driver: {src.driver}")
                log.info(f"Shapefile CRS: {src.crs}")
                log.info(f"Shapefile schema: {src.schema}")
                log.info(f"Shapefile feature count: {len(src)}")
                
                bounds = src.bounds
                log.info(f"Raw bounds: {bounds}")
                crs = src.crs
                
                # Transform to WGS84 if needed
                if crs and crs != from_epsg(4326):
                    log.info(f"Transforming bounds from {crs} to WGS84")
                    bounds = self._transform_bounds(bounds, crs, from_epsg(4326))
                    log.info(f"Transformed bounds: {bounds}")
                else:
                    log.info("Bounds already in WGS84 or no CRS info")
                
                result = self._bounds_to_geojson(bounds)
                log.info(f"Final GeoJSON result: {result}")
                return result
                
        except Exception as e:
            log.error(f"Error reading shapefile {file_path}: {str(e)}", exc_info=True)
            return None
    
    def _extract_raster_extent(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract extent from raster file (GeoTIFF, etc.)."""
        if not RASTERIO_AVAILABLE:
            return None
        
        try:
            with rasterio.open(file_path) as src:
                bounds = src.bounds
                crs = src.crs
                
                # Transform to WGS84 if needed
                if crs and crs.to_epsg() != 4326:
                    bounds = transform_bounds(crs, 'EPSG:4326', *bounds)
                
                return self._bounds_to_geojson(bounds)
                
        except Exception as e:
            log.debug(f"Error reading raster {file_path}: {str(e)}")
            return None
    
    def _extract_vector_extent(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract extent from vector files (KML, GeoPackage, GeoJSON)."""
        if not FIONA_AVAILABLE:
            return None
        
        try:
            with fiona.open(file_path) as src:
                bounds = src.bounds
                crs = src.crs
                
                # Transform to WGS84 if needed
                if crs and crs != from_epsg(4326):
                    bounds = self._transform_bounds(bounds, crs, from_epsg(4326))
                
                return self._bounds_to_geojson(bounds)
                
        except Exception as e:
            log.debug(f"Error reading vector file {file_path}: {str(e)}")
            return None
    
    def _extract_zip_shapefile_extent(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract extent from ZIP file containing Shapefile."""
        log.info(f"Extracting extent from ZIP shapefile: {file_path}")
        
        if not FIONA_AVAILABLE:
            log.info("Fiona not available for ZIP shapefile extraction")
            return None
            
        # Verify the file exists and has content
        if not os.path.exists(file_path):
            log.info(f"ZIP file not found: {file_path}")
            return None
            
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            log.info(f"ZIP file is empty (0 bytes): {file_path}")
            return None
            
        log.info(f"Processing ZIP file: {file_path}, size: {file_size} bytes")
        
        try:
            # Validate ZIP file
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_check:
                    zip_check.testzip()  # Test the integrity of the zip file
                    file_list = zip_check.namelist()
                    log.info(f"ZIP validation passed. Contains {len(file_list)} files.")
            except zipfile.BadZipFile as e:
                log.info(f"Invalid ZIP file format: {str(e)}")
                return None
            except Exception as e:
                log.info(f"Error testing ZIP file: {str(e)}")
                return None
                
            # Create temporary directory to extract ZIP contents
            with tempfile.TemporaryDirectory() as temp_dir:
                log.info(f"Created temporary directory: {temp_dir}")
                
                # Extract ZIP file
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    file_list = zip_ref.namelist()
                    log.info(f"ZIP contains {len(file_list)} files: {file_list}")
                    
                    # Check for common shapefile components
                    shp_files = [f for f in file_list if f.lower().endswith('.shp')]
                    shx_files = [f for f in file_list if f.lower().endswith('.shx')]
                    dbf_files = [f for f in file_list if f.lower().endswith('.dbf')]
                    
                    log.info(f"Found shapefile components - SHP: {len(shp_files)}, SHX: {len(shx_files)}, DBF: {len(dbf_files)}")
                    
                    # Extract all files
                    zip_ref.extractall(temp_dir)
                    log.info(f"Extracted ZIP contents to: {temp_dir}")
                
                # Look for .shp file in extracted contents
                shp_file = None
                for root, dirs, files in os.walk(temp_dir):
                    log.info(f"Checking directory: {root}, files: {files}")
                    for file in files:
                        if file.lower().endswith('.shp'):
                            shp_file = os.path.join(root, file)
                            log.info(f"Found shapefile: {shp_file}")
                            break
                    if shp_file:
                        break
                
                if not shp_file:
                    log.info(f"No .shp file found in ZIP: {file_path}")
                    return None
                
                # Check if shapefile exists and is readable
                if not os.path.exists(shp_file):
                    log.info(f"Shapefile path does not exist: {shp_file}")
                    return None
                
                shp_size = os.path.getsize(shp_file)
                log.info(f"Shapefile size: {shp_size} bytes")
                
                # Ensure the associated .shx and .dbf files exist
                shp_base = os.path.splitext(shp_file)[0]
                shx_file = shp_base + '.shx'
                dbf_file = shp_base + '.dbf'
                
                if not os.path.exists(shx_file):
                    log.info(f".shx file missing: {shx_file}")
                    return None
                    
                if not os.path.exists(dbf_file):
                    log.info(f".dbf file missing: {dbf_file}")
                    return None
                
                # Extract extent from the shapefile
                log.info(f"Extracting extent from shapefile: {shp_file}")
                result = self._extract_shapefile_extent(shp_file)
                log.info(f"Shapefile extent extraction result: {result}")
                return result
                
        except zipfile.BadZipFile as e:
            log.info(f"Bad ZIP file {file_path}: {str(e)}")
            return None
        except Exception as e:
            log.error(f"Error reading ZIP shapefile {file_path}: {str(e)}", exc_info=True)
            return None
    
    def _transform_bounds(self, bounds: Tuple[float, float, float, float], 
                         src_crs: Any, dst_crs: Any) -> Tuple[float, float, float, float]:
        """Transform bounds from source CRS to destination CRS."""
        if not PYPROJ_AVAILABLE:
            log.debug("PyProj not available - returning original bounds")
            return bounds
        
        try:
            transformer = Transformer.from_crs(src_crs, dst_crs, always_xy=True)
            min_x, min_y = transformer.transform(bounds[0], bounds[1])
            max_x, max_y = transformer.transform(bounds[2], bounds[3])
            return (min_x, min_y, max_x, max_y)
        except Exception as e:
            log.debug(f"Error transforming bounds: {str(e)}")
            return bounds
    
    def _bounds_to_geojson(self, bounds: Tuple[float, float, float, float]) -> Dict[str, Any]:
        """Convert bounds tuple to GeoJSON Polygon."""
        min_x, min_y, max_x, max_y = bounds
        
        # Create a polygon from the bounds
        coordinates = [[
            [min_x, min_y],  # Bottom-left
            [max_x, min_y],  # Bottom-right
            [max_x, max_y],  # Top-right
            [min_x, max_y],  # Top-left
            [min_x, min_y]   # Close the polygon
        ]]
        
        return {
            "type": "Polygon",
            "coordinates": coordinates
        }
    
    def extract_extent_from_upload(self, upload_file) -> Optional[Dict[str, Any]]:
        """
        Extract extent from an uploaded file.
        
        Args:
            upload_file: File upload object (from request)
            
        Returns:
            GeoJSON Polygon representing the extent, or None if extraction fails
        """
        if not hasattr(upload_file, 'filename') or not upload_file.filename:
            log.info("No filename in upload file")
            return None

        log.info(f"Processing upload file: {upload_file.filename}")
        
        # Get file extension
        ext = self._get_file_extension(upload_file.filename)
        log.info(f"File extension: {ext}")
        
        # Check if we can extract extent from this file type
        if not self.can_extract_extent(upload_file.filename, trust_extension=True):
            log.info(f"Cannot extract extent from file type: {ext}")
            return None

        # Create temporary file
        try:
            # Get file content
            upload_file.seek(0)
            file_content = upload_file.read()
            
            if not file_content:
                log.info("Upload file is empty")
                return None
            
            log.info(f"File content size: {len(file_content)} bytes")
            
            # Create temporary file with proper extension
            suffix = f".{ext}" if ext else ""
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
                tmp_file.write(file_content)
                tmp_file.flush()
                
                log.info(f"Created temporary file: {tmp_file.name}")
                
                # Extract extent
                extent = self.extract_extent(tmp_file.name, trust_extension=True)
                
                # Clean up
                try:
                    os.unlink(tmp_file.name)
                    log.info(f"Cleaned up temporary file: {tmp_file.name}")
                except OSError as e:
                    log.warning(f"Could not delete temporary file {tmp_file.name}: {e}")
                
                return extent
                
        except Exception as e:
            log.error(f"Error processing upload {upload_file.filename}: {str(e)}", exc_info=True)
            return None

    def extract_extent_from_url(self, url: str, file_format: str = None) -> Optional[Dict[str, Any]]:
        """
        Extract extent from a file accessible via URL.
        
        Args:
            url: URL to the file
            file_format: Optional format hint (e.g., 'shp', 'zip', 'tif')
            
        Returns:
            GeoJSON Polygon representing the extent, or None if extraction fails
        """
        if not url:
            log.info("No URL provided")
            return None

        log.info(f"Processing file from URL: {url}")
        log.info(f"Format hint: {file_format}")
        
        try:
            import urllib.request
            import urllib.error
            
            # Get file extension from URL or use format hint
            url_path = url.split('?')[0].split('#')[0]  # Remove query params
            ext = self._get_file_extension(url_path)
            
            # If we have a format hint and no extension, use the hint
            if not ext and file_format:
                ext = file_format.lower()
            
            log.info(f"Detected extension: {ext}")
            
            # For ZIP files or SHP format, always try to process
            # as they might contain shapefiles
            should_process = False
            if ext == 'zip' or file_format == 'shp':
                should_process = True
                log.info("Processing as potential shapefile (ZIP or SHP format)")
            elif self.can_extract_extent(f"dummy.{ext}", trust_extension=True):
                should_process = True
                log.info(f"Processing as supported format: {ext}")
            
            if not should_process:
                log.info(f"Cannot extract extent from format: {ext}")
                return None

            # Download file to temporary location
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp_file:
                log.info(f"Downloading file to: {tmp_file.name}")
                
                try:
                    # Add headers to avoid blocking
                    req = urllib.request.Request(url)
                    req.add_header('User-Agent', 'CKAN-SchemingDCAT-SpatialExtractor/1.0')
                    
                    with urllib.request.urlopen(req, timeout=30) as response:
                        # Read file in chunks
                        chunk_size = 8192
                        total_size = 0
                        while True:
                            chunk = response.read(chunk_size)
                            if not chunk:
                                break
                            tmp_file.write(chunk)
                            total_size += len(chunk)
                            # Limit file size to 100MB
                            if total_size > 100 * 1024 * 1024:
                                log.warning("File too large, aborting download")
                                return None
                    
                    tmp_file.flush()
                    log.info(f"Downloaded {total_size} bytes")
                    
                    if total_size == 0:
                        log.info("Downloaded file is empty")
                        return None
                    
                    # Extract extent from downloaded file
                    extent = self.extract_extent(tmp_file.name, trust_extension=True)
                    
                    return extent
                    
                except urllib.error.URLError as e:
                    log.error(f"Error downloading file from {url}: {str(e)}")
                    return None
                finally:
                    # Clean up
                    try:
                        os.unlink(tmp_file.name)
                        log.info(f"Cleaned up temporary file: {tmp_file.name}")
                    except OSError as e:
                        log.warning(f"Could not delete temporary file {tmp_file.name}: {e}")
                        
        except Exception as e:
            log.error(f"Error processing URL {url}: {str(e)}", exc_info=True)
            return None

    def extract_extent_from_resource(self, resource_url, resource_format=None) -> Optional[Dict[str, Any]]:
        """
        Extract spatial extent from a resource based on its URL and format.
        This method is designed for post-upload processing of files already stored in cloud storage.
        
        Args:
            resource_url: URL of the resource (e.g., Azure blob storage URL)
            resource_format: Format of the resource (e.g., 'SHP', 'ZIP', 'TIF')
            
        Returns:
            GeoJSON Polygon representing the extent, or None if extraction fails
        """
        log.info(f"Processing resource: {resource_url} (format: {resource_format})")
        
        # If format indicates shapefile, try to extract regardless of file extension
        if resource_format and resource_format.upper() in ['SHP', 'SHAPEFILE']:
            log.info("Resource format indicates shapefile - attempting extraction")
            return self.extract_extent_from_url(resource_url)
        
        # Check if it's a potential spatial file
        if self._is_potential_spatial_file(resource_url):
            log.info("Resource appears to be spatial file - attempting extraction")
            return self.extract_extent_from_url(resource_url)
        
        # If it's a ZIP file and format is not specified, check contents
        ext = self._get_file_extension(resource_url)
        if ext == 'zip':
            log.info("Resource is ZIP file - checking for spatial content")
            return self.extract_extent_from_url(resource_url)
        
        log.info(f"Resource does not appear to be spatial: {resource_url} (format: {resource_format})")
        return None


# Global instance
extent_extractor = SpatialExtentExtractor()


def extract_spatial_extent(file_path: str, trust_extension: bool = False) -> Optional[str]:
    """
    Convenience function to extract spatial extent from a file.
    
    Args:
        file_path: Path to the geospatial file
        trust_extension: If True, trust the file extension without validating content
        
    Returns:
        JSON string of the extent geometry, or None if extraction fails
    """
    extent = extent_extractor.extract_extent(file_path, trust_extension=trust_extension)
    return json.dumps(extent) if extent else None


def can_extract_spatial_extent(file_path: str, trust_extension: bool = False) -> bool:
    """
    Check if spatial extent can be extracted from the given file.
    
    Args:
        file_path: Path to the file
        trust_extension: If True, trust the file extension without validating content
        
    Returns:
        True if extent extraction is supported for this file type
    """
    return extent_extractor.can_extract_extent(file_path, trust_extension=trust_extension)


def get_spatial_system_status() -> Dict[str, Any]:
    """
    Get the status of the spatial extent extraction system.
    
    Returns:
        Dictionary with system status information
    """
    return {
        'available': any(extent_extractor.available_handlers.values()),
        'handlers': extent_extractor.available_handlers.copy(),
        'supported_extensions': list(extent_extractor.SUPPORTED_EXTENSIONS.keys()),
        'dependencies': {
            'fiona': FIONA_AVAILABLE,
            'rasterio': RASTERIO_AVAILABLE,
            'pyproj': PYPROJ_AVAILABLE,
            'pandas': PANDAS_AVAILABLE,
            'pdf': PDF_AVAILABLE,
            'excel': EXCEL_AVAILABLE
        },
        'api_safe': True,  # This system doesn't interfere with CKAN API
        'mode': 'frontend_only'  # Only works through web interface
    }


def analyze_file_comprehensive(file_path: str, trust_extension: bool = False) -> Dict[str, Any]:
    """
    Comprehensive analysis of a file to extract all available metadata.
    
    Args:
        file_path: Path to the file to analyze
        trust_extension: Whether to trust file extension without validation
        
    Returns:
        Dictionary with all extracted metadata including spatial, technical,
        and content-specific information
    """
    try:
        analyzer = FileAnalyzer()
        return analyzer.analyze_file(file_path, trust_extension)
    except Exception as e:
        log.error(f"Error in comprehensive file analysis: {str(e)}", exc_info=True)
        return {}


def analyze_upload_file(upload_file) -> Dict[str, Any]:
    """
    Analyze an uploaded file and extract comprehensive metadata.
    
    Args:
        upload_file: File upload object (from request)
        
    Returns:
        Dictionary with all extracted metadata
    """
    if not hasattr(upload_file, 'filename') or not upload_file.filename:
        log.info("No filename in upload file")
        return {}

    log.info(f"Processing upload file: {upload_file.filename}")
    
    try:
        # Get file content
        upload_file.seek(0)
        file_content = upload_file.read()
        
        if not file_content:
            log.info("Upload file is empty")
            return {}
        
        log.info(f"File content size: {len(file_content)} bytes")
        
        # Get file extension
        analyzer = FileAnalyzer()
        ext = analyzer.spatial_extractor._get_file_extension(upload_file.filename)
        
        # Create temporary file with proper extension
        suffix = f".{ext}" if ext else ""
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            tmp_file.write(file_content)
            tmp_file.flush()
            
            log.info(f"Created temporary file: {tmp_file.name}")
            
            # Analyze file
            result = analyzer.analyze_file(tmp_file.name, trust_extension=True)
            
            # Clean up
            try:
                os.unlink(tmp_file.name)
            except Exception:
                pass  # Ignore cleanup errors
            
            return result
            
    except Exception as e:
        log.error(f"Error analyzing upload file: {str(e)}", exc_info=True)
        return {}


def get_file_analysis_capabilities() -> Dict[str, Any]:
    """
    Get information about file analysis capabilities.
    
    Returns:
        Dictionary with available analysis features
    """
    return {
        'spatial_analysis': {
            'available': FIONA_AVAILABLE or RASTERIO_AVAILABLE,
            'supported_formats': ['shp', 'zip', 'tif', 'tiff', 'kml', 'gpkg', 'geojson', 'json']
        },
        'document_analysis': {
            'pdf': PDF_AVAILABLE,
            'excel': EXCEL_AVAILABLE,
            'csv': PANDAS_AVAILABLE,
            'text': True,
            'json': True
        },
        'data_analysis': {
            'statistics': PANDAS_AVAILABLE,
            'data_profiling': PANDAS_AVAILABLE
        },
        'technical_analysis': {
            'file_info': True,
            'checksums': True,
            'content_type_detection': True
        }
    }


class FileAnalyzer:
    """Comprehensive file analyzer for extracting metadata from various file types."""
    
    def __init__(self):
        self.spatial_extractor = SpatialExtentExtractor()
        
    def analyze_file(self, file_path: str, trust_extension: bool = False) -> Dict[str, Any]:
        """
        Analyze a file and extract comprehensive metadata.
        
        Args:
            file_path: Path to the file to analyze
            trust_extension: Whether to trust file extension without validation
            
        Returns:
            Dictionary containing all extracted metadata
        """
        try:
            result = {
                # Basic file information
                'file_size_bytes': None,
                'file_created_date': None,
                'file_modified_date': None,
                'content_type_detected': None,
                'file_integrity': {},
                
                # Spatial information (if applicable)
                'spatial_extent': None,
                'spatial_crs': None,
                'spatial_resolution': None,
                'feature_count': None,
                'geometry_type': None,
                
                # Data information
                'data_fields': None,
                'data_statistics': None,
                'data_domains': None,
                
                # Geographic information
                'geographic_coverage': None,
                'administrative_boundaries': None,
                
                # Temporal information
                'data_temporal_coverage': None,
                
                # Technical information
                'compression_info': None,
                'format_version': None,
                
                # Non-spatial file information
                'document_pages': None,
                'spreadsheet_sheets': None,
                'text_content_info': None
            }
            
            if not os.path.exists(file_path):
                return result
            
            # Extract basic file information
            self._extract_basic_file_info(file_path, result)
            
            # Detect content type
            self._detect_content_type(file_path, result)
            
            # Try spatial analysis first
            if self.spatial_extractor._is_potential_spatial_file(file_path):
                self._extract_spatial_info(file_path, result, trust_extension)
            
            # Extract content-specific information
            self._extract_content_specific_info(file_path, result)
            
            return result
            
        except Exception as e:
            log.error(f"Error analyzing file {file_path}: {str(e)}", exc_info=True)
            return result
    
    def _extract_basic_file_info(self, file_path: str, result: Dict[str, Any]):
        """Extract basic file system information."""
        try:
            stat = os.stat(file_path)
            result['file_size_bytes'] = stat.st_size
            result['file_created_date'] = datetime.fromtimestamp(stat.st_ctime).isoformat()
            result['file_modified_date'] = datetime.fromtimestamp(stat.st_mtime).isoformat()
            
            # Calculate checksums for integrity
            with open(file_path, 'rb') as f:
                content = f.read()
                result['file_integrity'] = {
                    'md5': hashlib.md5(content).hexdigest(),
                    'sha256': hashlib.sha256(content).hexdigest()
                }
        except Exception as e:
            log.debug(f"Error extracting basic file info: {str(e)}")
    
    def _detect_content_type(self, file_path: str, result: Dict[str, Any]):
        """Detect the content type of the file."""
        try:
            # Use mimetypes to guess
            mime_type, _ = mimetypes.guess_type(file_path)
            if mime_type:
                result['content_type_detected'] = mime_type
            else:
                # Fallback to extension-based detection
                ext = self.spatial_extractor._get_file_extension(file_path).lower()
                ext_to_mime = {
                    'pdf': 'application/pdf',
                    'csv': 'text/csv',
                    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'xls': 'application/vnd.ms-excel',
                    'txt': 'text/plain',
                    'json': 'application/json',
                    'xml': 'application/xml',
                    'zip': 'application/zip'
                }
                result['content_type_detected'] = ext_to_mime.get(ext, 'application/octet-stream')
        except Exception as e:
            log.debug(f"Error detecting content type: {str(e)}")
    
    def _extract_spatial_info(self, file_path: str, result: Dict[str, Any], trust_extension: bool = False):
        """Extract spatial information from geospatial files."""
        try:
            # Get spatial extent
            extent = self.spatial_extractor.extract_extent(file_path, trust_extension)
            if extent:
                result['spatial_extent'] = json.dumps(extent)
            
            # Extract additional spatial metadata
            ext = self.spatial_extractor._get_file_extension(file_path)
            format_type = self.spatial_extractor.SUPPORTED_EXTENSIONS.get(ext)
            
            if format_type == 'shapefile' and FIONA_AVAILABLE:
                self._extract_shapefile_metadata(file_path, result)
            elif format_type == 'geotiff' and RASTERIO_AVAILABLE:
                self._extract_raster_metadata(file_path, result)
            elif format_type in ['kml', 'geopackage', 'geojson'] and FIONA_AVAILABLE:
                self._extract_vector_metadata(file_path, result)
            elif format_type == 'zip_shapefile':
                self._extract_zip_shapefile_metadata(file_path, result)
                
        except Exception as e:
            log.debug(f"Error extracting spatial info: {str(e)}")
    
    def _extract_shapefile_metadata(self, file_path: str, result: Dict[str, Any]):
        """Extract metadata from shapefiles."""
        try:
            with fiona.open(file_path) as src:
                # CRS information
                if src.crs:
                    crs_info = str(src.crs)
                    if 'epsg' in src.crs:
                        crs_info = f"EPSG:{src.crs['init'].split(':')[1]}"
                    result['spatial_crs'] = crs_info
                
                # Feature count and geometry type
                result['feature_count'] = len(src)
                if len(src) > 0:
                    first_feature = next(iter(src))
                    result['geometry_type'] = first_feature['geometry']['type']
                
                # Schema information
                schema = src.schema
                fields_info = {}
                for field_name, field_type in schema['properties'].items():
                    fields_info[field_name] = field_type
                result['data_fields'] = json.dumps(fields_info)
                
                # Sample data for statistics (first 100 features)
                if PANDAS_AVAILABLE and len(src) > 0:
                    self._extract_vector_statistics(src, result)
                    
        except Exception as e:
            log.debug(f"Error extracting shapefile metadata: {str(e)}")
    
    def _extract_raster_metadata(self, file_path: str, result: Dict[str, Any]):
        """Extract metadata from raster files."""
        try:
            with rasterio.open(file_path) as src:
                # CRS information
                if src.crs:
                    result['spatial_crs'] = f"EPSG:{src.crs.to_epsg()}" if src.crs.to_epsg() else str(src.crs)
                
                # Resolution information
                transform = src.transform
                pixel_size_x = abs(transform[0])
                pixel_size_y = abs(transform[4])
                result['spatial_resolution'] = f"{pixel_size_x:.2f}m x {pixel_size_y:.2f}m"
                
                # Raster information
                result['data_fields'] = json.dumps({
                    'bands': src.count,
                    'width': src.width,
                    'height': src.height,
                    'dtype': str(src.dtypes[0]) if src.dtypes else None
                })
                
                # Basic statistics
                stats = {}
                for i in range(1, min(src.count + 1, 4)):  # Max 3 bands for performance
                    try:
                        band_data = src.read(i, masked=True)
                        stats[f'band_{i}'] = {
                            'min': float(band_data.min()),
                            'max': float(band_data.max()),
                            'mean': float(band_data.mean()),
                            'nodata_count': int(band_data.mask.sum())
                        }
                    except Exception:
                        continue
                
                if stats:
                    result['data_statistics'] = json.dumps(stats)
                    
        except Exception as e:
            log.debug(f"Error extracting raster metadata: {str(e)}")
    
    def _extract_vector_metadata(self, file_path: str, result: Dict[str, Any]):
        """Extract metadata from vector files (KML, GeoPackage, GeoJSON)."""
        try:
            with fiona.open(file_path) as src:
                # Similar to shapefile but for other vector formats
                if src.crs:
                    result['spatial_crs'] = str(src.crs)
                
                result['feature_count'] = len(src)
                if len(src) > 0:
                    first_feature = next(iter(src))
                    result['geometry_type'] = first_feature['geometry']['type']
                
                # Schema information
                schema = src.schema
                fields_info = {}
                for field_name, field_type in schema['properties'].items():
                    fields_info[field_name] = field_type
                result['data_fields'] = json.dumps(fields_info)
                
                if PANDAS_AVAILABLE and len(src) > 0:
                    self._extract_vector_statistics(src, result)
                    
        except Exception as e:
            log.debug(f"Error extracting vector metadata: {str(e)}")
    
    def _extract_zip_shapefile_metadata(self, file_path: str, result: Dict[str, Any]):
        """Extract metadata from ZIP files containing shapefiles."""
        try:
            import zipfile
            import tempfile
            
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # Find .shp file in the ZIP
                shp_files = [f for f in zip_ref.namelist() if f.endswith('.shp')]
                if not shp_files:
                    return
                
                # Extract to temporary directory
                with tempfile.TemporaryDirectory() as temp_dir:
                    zip_ref.extractall(temp_dir)
                    shp_path = os.path.join(temp_dir, shp_files[0])
                    self._extract_shapefile_metadata(shp_path, result)
                    
        except Exception as e:
            log.debug(f"Error extracting ZIP shapefile metadata: {str(e)}")
    
    def _extract_vector_statistics(self, src, result: Dict[str, Any]):
        """Extract statistics from vector data."""
        try:
            # Sample data for analysis (first 100 features)
            features = []
            for i, feature in enumerate(src):
                if i >= 100:  # Limit for performance
                    break
                features.append(feature['properties'])
            
            if not features:
                return
            
            # Convert to DataFrame for analysis
            df = pd.DataFrame(features)
            
            # Basic statistics
            stats = {}
            domains = {}
            
            for column in df.columns:
                if df[column].dtype in ['object', 'string']:
                    # Categorical data
                    unique_values = df[column].value_counts().head(10).to_dict()
                    domains[column] = unique_values
                    stats[column] = {
                        'type': 'categorical',
                        'unique_count': df[column].nunique(),
                        'null_count': df[column].isnull().sum()
                    }
                else:
                    # Numerical data
                    stats[column] = {
                        'type': 'numerical',
                        'min': df[column].min(),
                        'max': df[column].max(),
                        'mean': df[column].mean(),
                        'null_count': df[column].isnull().sum()
                    }
            
            result['data_statistics'] = json.dumps(stats)
            if domains:
                result['data_domains'] = json.dumps(domains)
                
        except Exception as e:
            log.debug(f"Error extracting vector statistics: {str(e)}")
    
    def _extract_content_specific_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information specific to non-spatial file types."""
        try:
            ext = self.spatial_extractor._get_file_extension(file_path).lower()
            
            if ext == 'pdf' and PDF_AVAILABLE:
                self._extract_pdf_info(file_path, result)
            elif ext in ['xlsx', 'xls'] and EXCEL_AVAILABLE:
                self._extract_excel_info(file_path, result)
            elif ext == 'csv' and PANDAS_AVAILABLE:
                self._extract_csv_info(file_path, result)
            elif ext in ['txt', 'md']:
                self._extract_text_info(file_path, result)
            elif ext == 'json':
                self._extract_json_info(file_path, result)
                
        except Exception as e:
            log.debug(f"Error extracting content-specific info: {str(e)}")
    
    def _extract_pdf_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information from PDF files."""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                result['document_pages'] = len(pdf_reader.pages)
                
                # Extract text content info
                text_length = 0
                for page in pdf_reader.pages[:5]:  # First 5 pages for performance
                    text_length += len(page.extract_text())
                
                result['text_content_info'] = json.dumps({
                    'estimated_characters': text_length * (len(pdf_reader.pages) / min(5, len(pdf_reader.pages))),
                    'pages_analyzed': min(5, len(pdf_reader.pages))
                })
                
        except Exception as e:
            log.debug(f"Error extracting PDF info: {str(e)}")
    
    def _extract_excel_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information from Excel files."""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            result['spreadsheet_sheets'] = len(workbook.sheetnames)
            
            # Analyze first sheet structure
            if workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                result['data_fields'] = json.dumps({
                    'sheets': workbook.sheetnames,
                    'first_sheet_dimensions': f"{max_row} rows x {max_col} columns"
                })
                
        except Exception as e:
            log.debug(f"Error extracting Excel info: {str(e)}")
    
    def _extract_csv_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information from CSV files."""
        try:
            # Read first few rows to analyze structure
            df = pd.read_csv(file_path, nrows=100)
            
            fields_info = {}
            stats = {}
            domains = {}
            
            for column in df.columns:
                fields_info[column] = str(df[column].dtype)
                
                if df[column].dtype in ['object', 'string']:
                    unique_values = df[column].value_counts().head(10).to_dict()
                    domains[column] = unique_values
                    stats[column] = {
                        'type': 'categorical',
                        'unique_count': df[column].nunique(),
                        'null_count': df[column].isnull().sum()
                    }
                else:
                    stats[column] = {
                        'type': 'numerical',
                        'min': df[column].min(),
                        'max': df[column].max(),
                        'mean': df[column].mean(),
                        'null_count': df[column].isnull().sum()
                    }
            
            result['data_fields'] = json.dumps(fields_info)
            result['data_statistics'] = json.dumps(stats)
            if domains:
                result['data_domains'] = json.dumps(domains)
                
        except Exception as e:
            log.debug(f"Error extracting CSV info: {str(e)}")
    
    def _extract_text_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information from text files."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(10000)  # First 10KB for analysis
                
            result['text_content_info'] = json.dumps({
                'character_count': len(content),
                'line_count': content.count('\n'),
                'word_count': len(content.split()),
                'encoding': 'utf-8'
            })
            
        except Exception as e:
            log.debug(f"Error extracting text info: {str(e)}")
    
    def _extract_json_info(self, file_path: str, result: Dict[str, Any]):
        """Extract information from JSON files."""
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
            
            result['text_content_info'] = json.dumps({
                'json_structure': type(data).__name__,
                'keys_count': len(data) if isinstance(data, dict) else None,
                'array_length': len(data) if isinstance(data, list) else None
            })
            
        except Exception as e:
            log.debug(f"Error extracting JSON info: {str(e)}")